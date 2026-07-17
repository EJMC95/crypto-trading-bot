#!/usr/bin/env python3
"""Build the fleet assessment workbook (2026-07-01)."""
import csv, os, datetime as dt
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = "/Users/eamonjuaomartins-carrick/Library/Application Support/Claude/local-agent-mode-sessions/f9b37507-d0ef-4e70-b1f2-919d1f04c42f/d91d4d39-ce69-4e05-a61c-491acc2e79f0/local_173eeaa3-3cd3-41a8-998f-50e0c7c7c35a/outputs"
XLSX = os.path.join(OUT, "MPG_Bot_Fleet_Assessment_2026-07-01.xlsx")

def load(name):
    p = os.path.join(HERE, name + ".csv")
    with open(p) as f:
        return list(csv.DictReader(f))

trades = load("bot_trades")
paper = load("paper_trades")
pnl = load("bot_pnl")

# ---- palette ----
NAVY="1F3864"; BLUE="2E5496"; LBLUE="D6E4F0"; GREY="F2F2F2"
GREEN="C6EFCE"; GREEND="006100"; RED="FFC7CE"; REDD="9C0006"
AMBER="FFEB9C"; AMBERD="7F6000"; WHITE="FFFFFF"
thin=Side(style="thin",color="BFBFBF")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws, row, headers, start=1, fill=NAVY):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=start+i,value=h)
        c.font=Font(bold=True,color=WHITE,size=10)
        c.fill=PatternFill("solid",fgColor=fill)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=BORDER
def title(ws,text,cells,size=15):
    ws.merge_cells(cells)
    c=ws[cells.split(":")[0]]
    c.value=text; c.font=Font(bold=True,color=WHITE,size=size)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center")
print("loaded", len(trades),"trades", len(paper),"paper", len(pnl),"pnl rows")

# ============ CURATED PER-BOT ASSESSMENT ============
# status: WIN / LOSS / FLAT / IDLE ; health: OK / FIX / KILL? / WATCH
FLEET = [
 # name, venue, kind, strategy, tf, pnl, wl, verdict, health
 ["alpaca-momentum","Alpaca (paper)","stocks","Momentum rank top-N (SMA trend+cross)","1d",
   3179.63,"n/a","STAR. +3.06%; 96k->107k. Rides tech momentum (AMD,CRWD,PANW,TSLA).","OK"],
 ["ikbr-stock-bot","IBKR (paper)","stocks","SPY/QQQ regime (SMA cross / sma_rsi)","1d",
   307.21,"n/a","SOUND, flat +0.12%. Holds 16 SPY+16 QQQ. Slow regime, working as designed.","OK"],
 ["listing-sniper","Multi-CEX (paper)","event","New-listing snipe + research gate","tick",
   272.89,"3W/2L","ONE winner (ANSEM +$325). Bought SOXL3L (3x leveraged token) -$50 - universe leak.","FIX"],
 ["trend-golden-cross","Kraken (paper)","crypto-spot","ImprovedStrategyV4 50/200 golden cross","1d",
   0.0,"0W/0L","SOUND, correctly holding. 0 closed is normal (~2-3 trades/yr). Riding uptrend.","OK"],
 ["hl-perps-rsi","Hyperliquid (paper)","perps","RSI(14) mean-revert, trend-gated 30/70","loop",
   0.0,"0/0","SOUND but near-idle. 30/70 + trend gate = rare signals. Under-participates.","WATCH"],
 ["swing-dip-buyer","Kraken (paper)","crypto-spot","SwingDipV1 RSI<35 + BB-lower in uptrend","1d",
   0.0,"0W/0L","SOUND but IDLE. No valid dip fired (backtest ~7-13 trades / 3yr). Not broken.","WATCH"],
 ["hl-momo-breakout","Hyperliquid (paper)","perps","Donchian 12/8 + 100EMA, shorts on, lev","loop",
   0.14,"n/a","FLAT roundtrip (969-1033->1000). 'Relaxed for frequency' added chop churn.","FIX"],
 ["triangular-arb","Kraken (paper)","scanner","Tri-arb depth scan","tick",
   0.0,"n/a","SCANNER, 0 fills (best depth -1.1%, below fee threshold). Correct = monitor.","OK"],
 ["cross-exchange-arb","Multi-CEX (paper)","scanner","Cross-exchange spread scan","tick",
   0.0,"n/a","SCANNER. Reports +$560 THEORETICAL spread (not executable). Exclude from P&L.","OK"],
 ["momo-breakout-4h","Kraken (paper)","crypto-spot","MomoBreakoutV1 Donchian 30/15 + 200EMA","4h",
   -0.70,"0W/1L","SOUND, tiny sample (1 trade TRX -1.7%). PF 1.8-1.9 in backtest. Needs time.","OK"],
 ["momo-breakout-alt","Kraken (paper)","crypto-spot","TrendMomoV1 20/50 SMA (sped to 4h)","4h",
   -11.33,"0W/2L","CONTRADICTION: loose entry (close>slowSMA) + strict cross-down exit -> SOL -9.5% giveback.","FIX"],
 ["intraday-daytrader-5m","Kraken (paper)","crypto-spot","DayTraderV5Gated 5m EMA/RSI + regime gate","5m",
   -18.85,"12W/107L","THE BLEEDER. 10% win. VolumePairList junk + tight ATR trail + 5m fee churn.","KILL?"],
]

wb=openpyxl.Workbook()

# =================== SHEET 1: FLEET SUMMARY ===================
ws=wb.active; ws.title="Fleet Summary"
title(ws,"MPG Trading-Bot Fleet — Month Assessment (all DRY-RUN / paper)  ·  data 2026-06-22→07-01","A1:I1")
ws["A2"]="Cumulative P&L since fleet reset. Crypto wallets $1,000 each; stocks are paper accounts. cross-exchange-arb P&L is theoretical (not executable)."
ws["A2"].font=Font(italic=True,size=9,color="555555"); ws.merge_cells("A2:I2")
cols=["Bot","Venue","Type","Strategy","TF","P&L $","W/L","Verdict","Health"]
hdr(ws,4,cols)
r=5
for b in FLEET:
    name,venue,kind,strat,tf,p,wl,verdict,health=b
    vals=[name,venue,kind,strat,tf,p,wl,verdict,health]
    for i,v in enumerate(vals):
        c=ws.cell(row=r,column=1+i,value=v); c.border=BORDER
        c.alignment=Alignment(vertical="center",wrap_text=(i in(3,7)))
        c.font=Font(size=9)
    ws.cell(row=r,column=1).font=Font(bold=True,size=9)
    pc=ws.cell(row=r,column=6); pc.number_format='#,##0.00;[Red]-#,##0.00'
    pc.font=Font(bold=True,size=9,color=GREEND if p>0 else (REDD if p<0 else "555555"))
    hc=ws.cell(row=r,column=9)
    fill={"OK":GREEN,"FIX":AMBER,"WATCH":LBLUE,"KILL?":RED}.get(health,WHITE)
    fdark={"OK":GREEND,"FIX":AMBERD,"WATCH":BLUE,"KILL?":REDD}.get(health,"000000")
    hc.fill=PatternFill("solid",fgColor=fill); hc.font=Font(bold=True,size=9,color=fdark)
    hc.alignment=Alignment(horizontal="center",vertical="center")
    if r%2==0:
        for i in range(1,10):
            if i not in(6,9): ws.cell(row=r,column=i).fill=PatternFill("solid",fgColor=GREY)
    r+=1
# totals
ws.cell(row=r,column=1,value="FLEET TOTAL (excl. theoretical arb)").font=Font(bold=True)
tot=sum(b[5] for b in FLEET if b[0]!="cross-exchange-arb")
tc=ws.cell(row=r,column=6,value=tot); tc.number_format='#,##0.00;[Red]-#,##0.00'
tc.font=Font(bold=True,color=GREEND if tot>0 else REDD)
ws.cell(row=r+1,column=1,value="  of which CRYPTO only").font=Font(italic=True,size=9)
cc=sum(b[5] for b in FLEET if b[2] in("crypto-spot","perps","event","scanner") and b[0]!="cross-exchange-arb")
cyc=ws.cell(row=r+1,column=6,value=cc); cyc.number_format='#,##0.00;[Red]-#,##0.00'
cyc.font=Font(italic=True,size=9,color=GREEND if cc>0 else REDD)
widths=[22,18,11,40,6,11,9,52,9]
for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width=w
ws.freeze_panes="A5"
print("sheet1 done")

# =================== SHEET 2: DIAGNOSIS & FIXES ===================
DIAG = [
 ["intraday-daytrader-5m","KILL? / REBUILD",
  "Won only via ROI target (5/5 ROI exits were wins).",
  "10% win rate (12/119). 95 of 119 exits are trailing-stop (ATR 1.5x on 5m ~ sub-1%) - cut before ROI. Dynamic VolumePairList pulled illiquid junk (M,CC,AKT,UAI,ZEREBRO). 5m on Kraken pays 0.52% round-trip fees - churn death. Every pair net-negative except SYN.",
  "Pick one: (A) RETIRE it - concept loses, its own header says daily trend (V4) is the real earner; reallocate. (B) FIX: swap VolumePairList->curated majors whitelist, widen ATR stop 1.5->3.0 so ROI can hit, keep regime gate. (C) Port your Donchian-breakout rework from ~/freqtrade. (D) Move to 15m/1h to cut fee churn. Rec: A or B."],
 ["momo-breakout-alt (TrendMomoV1)","FIX",
  "Trend-follow concept is robustness-checked (walk-forward OK).",
  "Internal contradiction: entry LOOSENED to (fast>slow OR close>slowSMA) so it enters on weak signals, but EXIT still requires strict crossed_below(fast,slow). Enters early, exits late -> holds losers down (SOL -9.5% giveback). Also -15% stop + comments still say 'daily' after 4h speed-up.",
  "Restore symmetry: entry = fast>slow only (drop the OR close>slow clause), OR add a faster protective exit (close<slowSMA). Tighten stop to -8% on 4h. This removes the enter-early/exit-late trap."],
 ["listing-sniper","FIX",
  "Research gate + scale-out worked on ANSEM: partials +$73/+$51 then +$200 TP = +$325.",
  "Whole month P&L is ONE listing (luck/concentration). SKIP_BASES filters only stables/fiat - NOT leveraged tokens, so it bought SOXL3L (3x token) -> -$50 stop. KGM -$2 trail.",
  "Add leveraged-token exclusion to base_skipped (regex: bases ending 3L/3S/5L/5S or containing UP/DOWN/BULL/BEAR). Keep gate. Treat P&L as small-sample - keep observing, don't scale."],
 ["hl-momo-breakout","FIX",
  "Donchian breakout port is a sound concept (mirrors MomoBreakoutV1 PF~1.8).",
  "Was 'relaxed for frequency': lookback 15->12 and ALLOW_SHORT=True. The validated backtest was long-only 15/8. Shorts + shorter window add whipsaw in chop -> equity roundtripped 969-1033 back to 1000 (no net edge).",
  "Revert to validated params: ENTRY_LOOKBACK 12->15, ALLOW_SHORT=False (or gate shorts behind a strong downtrend filter). Frequency isn't the goal - edge is."],
 ["hl-perps-rsi","WATCH / OPTIONAL",
  "Trend-gated mean-reversion (long dips above EMA, short rips below) is directionally sane; 3% stop added.",
  "RSI 30/70 + trend gate = signals almost never fire across 30 coins -> ~0 trades, no contribution. Equity dipped to 957 earlier then flat.",
  "Optional loosen to 35/65 to participate more, OR accept it as the 'only trade true extremes' sleeve and leave it. Low risk either way."],
 ["swing-dip-buyer","WATCH (leave)",
  "Correct discipline: only buys oversold dips (RSI<35 + below BB-lower) inside an uptrend - avoids falling knives (the V2 mistake).",
  "Zero trades in the month. Expected: backtest is ~7-13 trades over 3 YEARS. Not broken - just no qualifying setup yet.",
  "Leave as-is (diversifier). If you want more participation, relax RSI 35->40 or make the dip filter an OR - but this raises knife-catch risk. Rec: leave."],
 ["trend-golden-cross","OK (leave)",
  "Best-designed crypto bot: daily 50/200 golden cross; backtest +171% vs +155% hold, half the drawdown.",
  "0 closed trades looks idle but is correct - it enters and HOLDS the uptrend (currently long, small unrealized -1%).",
  "No change. Optionally widen the whitelist beyond BTC/ETH/SOL/BNB/XRP/TRX to catch more trends. This is your crypto core - keep it."],
 ["momo-breakout-4h","OK (leave)",
  "MomoBreakoutV1 Donchian 30/15 + 200EMA; backtest PF 1.8-1.9, ~40% win (normal for breakout).",
  "1 trade (TRX -1.7%) = low frequency, not a fault. 40% win means many small losers by design.",
  "No change. Give it a full quarter before judging. Don't panic at small losers."],
 ["triangular-arb","OK (leave)",
  "Working scanner: 848 cycles, correctly executes nothing (best depth -1.1%, below fee threshold).",
  "No P&L by design - it's a monitor for real tri-arb windows that rarely clear fees on one venue.",
  "Leave. It's costless surveillance. Consider alerting if best_depth ever turns positive."],
 ["cross-exchange-arb","OK (label clearly)",
  "Scans 342 pairs, finds 3-4% top-of-book spreads.",
  "The +$560 'P&L' is THEORETICAL spread - not executable (needs capital on both venues + instant transfer). Daily report already excludes it, good.",
  "Keep excluded from real totals. Useful as a market-dislocation signal only."],
 ["alpaca-momentum","OK (formalise risk)",
  "STAR performer +$3,180 (+3.06%). Clean momentum: price>trendSMA & fast>slow, rank by momentum, hold top-N.",
  "No per-position stop visible; gains are unrealised (7 open). A momentum unwind could give back fast.",
  "Add a trailing/ATR stop or a trend-exit per name to lock gains. Otherwise leave - it's the fleet's best."],
 ["ikbr-stock-bot","OK (leave)",
  "SPY/QQQ regime, single source of truth shared by backtest+live (trades exactly what was tested).",
  "Flat +0.12%; slow by design. Working correctly.",
  "No change. This is a low-vol ballast sleeve."],
]
ws=wb.create_sheet("Diagnosis & Fixes")
title(ws,"Per-Bot Diagnosis — what won, what failed/contradicts, and the fix (with options)","A1:E1")
hdr(ws,3,["Bot","Call","What WON","What FAILED / contradicts","Recommended fix & options"])
r=3
for d in DIAG:
    r+=1
    for i,v in enumerate(d):
        c=ws.cell(row=r,column=1+i,value=v); c.border=BORDER
        c.alignment=Alignment(vertical="top",wrap_text=True); c.font=Font(size=9)
    ws.cell(row=r,column=1).font=Font(bold=True,size=9)
    call=d[1]; cc=ws.cell(row=r,column=2)
    fill=AMBER if call.startswith("FIX") else (RED if "KILL" in call else (GREEN if call.startswith("OK") else LBLUE))
    cc.fill=PatternFill("solid",fgColor=fill); cc.font=Font(bold=True,size=9)
for i,w in enumerate([22,16,34,52,60]): ws.column_dimensions[get_column_letter(i+1)].width=w
ws.freeze_panes="A4"
print("sheet2 done")

# =================== SHEET 3: DAY-TRADER DEEP DIVE ===================
dt5=[t for t in trades if t["bot"]=="intraday-daytrader-5m"]
def fnum(x):
    try: return float(x)
    except: return 0.0
# by exit reason
byreason=defaultdict(lambda:[0,0,0.0])
for t in dt5:
    k=t["exit_reason"] or "?"; byreason[k][0]+=1
    if fnum(t["profit_abs"])>0: byreason[k][1]+=1
    byreason[k][2]+=fnum(t["profit_abs"])
# by pair
bypair=defaultdict(lambda:[0,0,0.0])
for t in dt5:
    k=t["pair"]; bypair[k][0]+=1
    if fnum(t["profit_abs"])>0: bypair[k][1]+=1
    bypair[k][2]+=fnum(t["profit_abs"])

ws=wb.create_sheet("DayTrader Deep-Dive")
title(ws,"intraday-daytrader-5m (DayTraderV5Gated) — 119 trades, -$18.85, 12W/107L (10% win)","A1:D1")
ws["A3"]="Exit reason (the core problem: 95/119 exits are the tight ATR trailing-stop, almost all losers)"
ws["A3"].font=Font(bold=True,color=NAVY)
hdr(ws,4,["Exit reason","Trades","Wins","Total P&L $"])
r=4
for k,v in sorted(byreason.items(),key=lambda kv:-kv[1][0]):
    r+=1
    ws.cell(row=r,column=1,value=k).border=BORDER
    ws.cell(row=r,column=2,value=v[0]).border=BORDER
    ws.cell(row=r,column=3,value=v[1]).border=BORDER
    pc=ws.cell(row=r,column=4,value=round(v[2],2)); pc.border=BORDER
    pc.number_format='#,##0.00;[Red]-#,##0.00'
    pc.font=Font(color=GREEND if v[2]>0 else REDD)
r+=2
ws.cell(row=r,column=1,value="P&L by pair (VolumePairList pulled illiquid junk; only SYN net-positive)").font=Font(bold=True,color=NAVY)
r+=1; hdr(ws,r,["Pair","Trades","Wins","Total P&L $"]); hrow=r
for k,v in sorted(bypair.items(),key=lambda kv:kv[1][2]):
    r+=1
    ws.cell(row=r,column=1,value=k).border=BORDER
    ws.cell(row=r,column=2,value=v[0]).border=BORDER
    ws.cell(row=r,column=3,value=v[1]).border=BORDER
    pc=ws.cell(row=r,column=4,value=round(v[2],2)); pc.border=BORDER
    pc.number_format='#,##0.00;[Red]-#,##0.00'
    pc.font=Font(color=GREEND if v[2]>0 else REDD)
for i,w in enumerate([20,10,8,12]): ws.column_dimensions[get_column_letter(i+1)].width=w
print("sheet3 done")

# =================== SHEET 4: ALL TRADES ===================
ws=wb.create_sheet("All Trades (raw)")
title(ws,"All closed trades pulled from Railway Postgres (bot_trades + paper_trades)","A1:I1")
cols=["Bot","Pair","Open (UTC)","Close (UTC)","Profit %","Profit $","Exit reason","Dur (min)","Source"]
hdr(ws,3,cols); r=3
allrows=[]
for t in trades:
    allrows.append([t["bot"],t["pair"],t["open_ts"][:16],(t["close_ts"] or "")[:16],
        round(fnum(t["profit_ratio"])*100,2),round(fnum(t["profit_abs"]),3),
        t["exit_reason"],fnum(t["duration_min"]),"bot_trades"])
for t in paper:
    allrows.append([t["bot"],t["pair"],t["opened_at"][:16],(t["closed_at"] or "")[:16],
        round(fnum(t["pnl_pct"])*100,2),round(fnum(t["pnl_abs"]),3),
        t["reason"],"","paper_trades"])
allrows.sort(key=lambda x:x[2])
for row in allrows:
    r+=1
    for i,v in enumerate(row):
        c=ws.cell(row=r,column=1+i,value=v); c.border=BORDER; c.font=Font(size=9)
    for col in(5,6):
        cc=ws.cell(row=r,column=col); cc.number_format='#,##0.00;[Red]-#,##0.00'
        cc.font=Font(size=9,color=GREEND if fnum(row[col-1])>0 else REDD)
for i,w in enumerate([22,14,17,17,10,11,20,10,12]): ws.column_dimensions[get_column_letter(i+1)].width=w
ws.freeze_panes="A4"
print("sheet4 done", len(allrows),"rows")

wb.save(XLSX)
print("SAVED:", XLSX)
